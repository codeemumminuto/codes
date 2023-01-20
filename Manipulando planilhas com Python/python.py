from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
lista = ['nome1','nome2','nome3','nome4','nome5','nome6','nome7','nome8','nome9','nome10',
        'nome11','nome12','nome13','nome14','nome15','nome16','nome17','nome18','nome19','nome20',
        'nome21','nome22','nome23','nome24','nome25','nome26','nome27','nome28','nome29','nome30',
        'nome31','nome32','nome33','nome34','nome35']

planilha1 = wb.active
planilha1.title = "teste"

i = 0
for coluna in range(1,6):
    for linha in range(1,8):
        letra = get_column_letter(coluna)
        planilha1[letra + str(linha)] = lista[i]
        i+=1
        
wb.save('Teste.xlsx')

